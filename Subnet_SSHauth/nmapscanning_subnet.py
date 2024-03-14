import subprocess
import re
from datetime import datetime
from openpyxl import Workbook
import ipaddress
import time


def is_valid_ip(subnet):
    try:
        ipaddress.ip_network(subnet, strict=False)
        return True
    except ValueError:
        print("Subnet Error!")
        time.sleep(5)
        return False

def nmap(subnet):
    cmd = "nmap " + subnet + " -p 22 --script=ssh-auth-methods.nse -Pn"
    print(cmd)
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    
    log_without_newline = re.sub(r'\\n', '', str(result))
    log_without_symbol = re.sub(r'[!@#$%^&*():|=]', '', log_without_newline)
    log_without_symbol = re.sub(r'Host', '', log_without_symbol)

    nmap = list()
    search_word ='Nmap'
    index1 = log_without_symbol.find(search_word)

    while index1 != -1:
        nmap.append(index1)
        index1 = log_without_symbol.find(search_word, index1+1)

    Text = []*len(nmap)

    for i in range(len(nmap)-1):
        Text.append(log_without_symbol[nmap[i]:nmap[i+1]])
    
    Text.pop(0)
    return Text

def nmap_label(Text):
    Text_split = [i.split() for i in Text]

    Hostlist = list()
    for j, k in enumerate(Text_split):
        index1 = Text_split[j].index('for')
        index2 = Text_split[j].index('is')
        Host = Text_split[j][index1+1:index2]
        if len(Host) == 1:
            Host.insert(0, '')
        Hostlist.append(Host)

    keyword = ['publickey', 'password', 'filtered']

    for j, m in enumerate(Text):
            for n, i in enumerate(keyword):
                if re.search(i, m):
                    Hostlist[j].append(1)
                else:
                    Hostlist[j].append(0)
    return Hostlist

def grouping(nuclist):       
        for j in nuclist:
            if j[4] == 1:    #unreachable
                j.append('SSH not Reachable/Filtered')
            elif j[2] == 1 and j[3] == 0:    #only publickey
                j.append('Compliant')
            elif j[2] == 1 and j[3] == 1:    #publickey + password
                j.append('Non-compliant')
            else:           #if nothing 
                j.append('Unable to get ssh_auth')
        return nuclist

def output_gen(nuclist, subnet):
    columns = ['Hostname','Host_IP','PublicKey', 'Password', 'Unreachable', 'Output']
    subnet_new = [re.sub(r'/.*$', '', cidr) for cidr in subnet]
    workbook = Workbook()

    for index, k in enumerate(nuclist):
        sheet = workbook.create_sheet(title=f'{subnet_new[index]}')
        sheet.append(columns)

        for l in k:
            sheet.append(l)
        
    # Sort the nuclist based on the desired order
    # custom_order = {'Non-compliant': 1, 'Reachable but unable to detect method': 2, 'Not Reachable': 3, 'Compliant': 4}
    # nuclist = sorted(nuclist, key=lambda x: custom_order.get(x[-1], float('inf')))
    
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    excel_file_path = f'data/output/output_{timestamp}.xlsx'
    first_sheet = workbook.sheetnames[0]
    workbook.remove(workbook[first_sheet])
    workbook.save(excel_file_path)
    print("Successfully performed nmap_checking!")

def nmap_checking(subnet):
    subnet_valid=list()
    for i in subnet:
        subnet_valid.append(is_valid_ip(i))
    
    if sum(subnet_valid) == len(subnet_valid):
        subnet_out=list()
        for item in subnet:
            X = grouping(nmap_label(nmap(item)))
            subnet_out.append(X)
        return output_gen(subnet_out, subnet)
    else:
        return("Wrong subnet/IP")
    
if __name__ == "__main__":
    
    file_path='data/subnet.txt'
    with open(file_path, 'r') as file:
        data = [line.strip() for line in file]
    
    print("Checking for :", data) 
    nmap_checking(data)